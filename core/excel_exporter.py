# core/excel_exporter.py
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


HEADERS = [
    "#",
    "Status",
    "Priority",
    "Task Description",
    "Category",
    "Deadline",
    "Entry Date",
    "Memo",
]

STATUS_DONE_VALUES = {"[x]", "[done]", "done", "✅"}

HEADER_FILL = PatternFill("solid", fgColor="BA55D3")
HEADER_FONT = Font(bold=True, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin", color="E5D5EF"),
    right=Side(style="thin", color="E5D5EF"),
    top=Side(style="thin", color="E5D5EF"),
    bottom=Side(style="thin", color="E5D5EF"),
)

STATUS_FILLS = {
    "Done": PatternFill("solid", fgColor="D9EAD3"),
    "Open": PatternFill("solid", fgColor="FFF2CC"),
}

PRIORITY_FILLS = {
    "P0": PatternFill("solid", fgColor="F4CCCC"),
    "P1": PatternFill("solid", fgColor="FCE5CD"),
    "P2": PatternFill("solid", fgColor="FFF2CC"),
    "P3": PatternFill("solid", fgColor="D9EAD3"),
}

OVERDUE_FILL = PatternFill("solid", fgColor="F4CCCC")
TODAY_FILL = PatternFill("solid", fgColor="FCE5CD")
SOON_FILL = PatternFill("solid", fgColor="FFF2CC")
ALT_ROW_FILL = PatternFill("solid", fgColor="F8F1FB")
DONE_ROW_FILL = PatternFill("solid", fgColor="EAF4E2")


def _normalize_cell(value):
    if value is None:
        return ""
    return str(value).strip()


def _display_status(status):
    clean_status = _normalize_cell(status)
    if clean_status.lower() in STATUS_DONE_VALUES:
        return "Done"
    if clean_status:
        return "Open"
    return ""


def _parse_deadline(value):
    raw_value = _normalize_cell(value)
    if not raw_value or raw_value.lower() in {"none", "no deadline", "n/a", "-"}:
        return None

    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(raw_value, fmt).date()
        except ValueError:
            continue

    return None


def _apply_deadline_style(cell, status):
    deadline = _parse_deadline(cell.value)
    if deadline is None or status == "Done":
        return

    days_left = (deadline - date.today()).days

    if days_left < 0:
        cell.fill = OVERDUE_FILL
        cell.font = Font(bold=True, color="990000")
    elif days_left == 0:
        cell.fill = TODAY_FILL
        cell.font = Font(bold=True, color="7F3F00")
    elif days_left <= 3:
        cell.fill = SOON_FILL
        cell.font = Font(bold=True, color="7F6000")


def export_tasks_to_excel(tasks, output_path, sheet_name="NoBrainFog Tasks"):
    """
    Write NoBrainFog task dictionaries to an .xlsx workbook.

    Args:
        tasks: Iterable of task dictionaries from TodoHandler.get_tasks().
        output_path: Target .xlsx path.
        sheet_name: Excel worksheet name.

    Returns:
        pathlib.Path for the generated workbook.
    """
    output = Path(output_path).expanduser().resolve()
    output.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name[:31]

    worksheet.append(HEADERS)

    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = header_alignment
        cell.border = THIN_BORDER

    for task in tasks:
        worksheet.append([
            task.get("number", ""),
            _display_status(task.get("status", "")),
            _normalize_cell(task.get("priority", "")),
            _normalize_cell(task.get("task", "")),
            _normalize_cell(task.get("category", "")),
            _normalize_cell(task.get("deadline", "")),
            _normalize_cell(task.get("entry_date", "")),
            _normalize_cell(task.get("memo", "")),
        ])

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    column_widths = {
        "A": 6,
        "B": 12,
        "C": 12,
        "D": 52,
        "E": 18,
        "F": 16,
        "G": 16,
        "H": 52,
    }

    for column_letter, width in column_widths.items():
        worksheet.column_dimensions[column_letter].width = width

    for row_index in range(2, worksheet.max_row + 1):
        status = _normalize_cell(worksheet.cell(row=row_index, column=2).value)
        priority = _normalize_cell(worksheet.cell(row=row_index, column=3).value)
        base_fill = DONE_ROW_FILL if status == "Done" else ALT_ROW_FILL if row_index % 2 == 0 else None

        worksheet.row_dimensions[row_index].height = 34

        for cell in worksheet[row_index]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER
            if base_fill is not None:
                cell.fill = base_fill

        status_cell = worksheet.cell(row=row_index, column=2)
        if status in STATUS_FILLS:
            status_cell.fill = STATUS_FILLS[status]
            status_cell.font = Font(bold=True, color="274E13" if status == "Done" else "7F6000")
            status_cell.alignment = Alignment(horizontal="center", vertical="center")

        priority_cell = worksheet.cell(row=row_index, column=3)
        if priority in PRIORITY_FILLS:
            priority_cell.fill = PRIORITY_FILLS[priority]
            priority_cell.font = Font(bold=True)
            priority_cell.alignment = Alignment(horizontal="center", vertical="center")

        _apply_deadline_style(worksheet.cell(row=row_index, column=6), status)

    workbook.save(output)
    return output
